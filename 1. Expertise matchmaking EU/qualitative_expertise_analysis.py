#!/usr/bin/env python3
"""
Qualitative Deep Expertise Analysis - Advanced semantic matching with validation loops.
Analyzes expertise documents for core capabilities and themes, then validates alignment
with call descriptions using semantic similarity and thematic matching.

OPTIMIZED FOR SPEED: Uses embedding caching and batch processing.
"""

import pandas as pd
import re
from pathlib import Path
from collections import defaultdict
from difflib import SequenceMatcher
import numpy as np
from rapidfuzz import fuzz

# Import embedding model for semantic similarity (same as EU Project Matcher)
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity

# Global embedding model (lazy-loaded)
_embedding_model = None
_embedding_cache = {}  # Cache for precomputed embeddings {text_hash: embedding}

def get_embedding_model():
    """Lazy-load embedding model to avoid loading multiple times."""
    global _embedding_model
    if _embedding_model is None:
        _embedding_model = SentenceTransformer('all-MiniLM-L6-v2')
    return _embedding_model

def batch_encode_texts(texts):
    """
    Encode multiple texts at once (faster than individual encoding).
    Returns embeddings dict {text_hash: embedding}.
    """
    model = get_embedding_model()
    texts_clean = [str(t).strip() for t in texts if str(t).strip()]
    
    if not texts_clean:
        return {}
    
    # Batch encoding is more efficient
    embeddings_tensor = model.encode(texts_clean, convert_to_tensor=True)
    
    result = {}
    for text, embedding in zip(texts_clean, embeddings_tensor):
        text_hash = hash(text)
        result[text_hash] = embedding.cpu().numpy()
    
    return result

def compute_semantic_similarity_cached(text1, text2):
    """
    Compute semantic similarity using cached embeddings (OPTIMIZED).
    Uses precomputed embeddings when available to avoid redundant computation.
    Falls back to on-demand encoding for new texts.
    
    Returns score between 0 and 1 (normalized).
    """
    if not text1 or not text2:
        return 0.0
    
    try:
        text1_clean = str(text1).strip()
        text2_clean = str(text2).strip()
        
        text1_hash = hash(text1_clean)
        text2_hash = hash(text2_clean)
        
        # Check cache first, encode if not cached
        if text1_hash not in _embedding_cache:
            embedding1_tensor = get_embedding_model().encode(text1_clean, convert_to_tensor=True)
            _embedding_cache[text1_hash] = embedding1_tensor.cpu().numpy()
        else:
            embedding1_tensor = _embedding_cache[text1_hash]
        
        if text2_hash not in _embedding_cache:
            embedding2_tensor = get_embedding_model().encode(text2_clean, convert_to_tensor=True)
            _embedding_cache[text2_hash] = embedding2_tensor.cpu().numpy()
        else:
            embedding2_tensor = _embedding_cache[text2_hash]
        
        # Compute cosine similarity from cached embeddings
        embedding1 = _embedding_cache[text1_hash]
        embedding2 = _embedding_cache[text2_hash]
        
        similarity = float(cosine_similarity([embedding1], [embedding2])[0][0])
        return min(1.0, max(0.0, similarity))  # Clamp to [0, 1]
        
    except Exception as e:
        print(f"[WARNING] Semantic similarity failed: {e}")
        # Fallback to fuzzy matching if embedding fails
        text1 = str(text1).lower().strip()
        text2 = str(text2).lower().strip()
        
        token_set = fuzz.token_set_ratio(text1, text2) / 100.0
        partial = fuzz.partial_ratio(text1, text2) / 100.0
        simple = fuzz.ratio(text1, text2) / 100.0
        
        combined = (token_set * 0.5) + (partial * 0.3) + (simple * 0.2)
        return min(1.0, max(0.0, combined))

# Keep old name as alias for compatibility
def compute_semantic_similarity(text1, text2):
    """Alias for cached version."""
    return compute_semantic_similarity_cached(text1, text2)


def segment_expertise_into_topics(expertise_info):
    """
    Split expertise into distinct research topics/lines.
    Each topic = section of text + associated keywords.
    
    Strategy: Use themes and capabilities as topic boundaries.
    """
    topics = []
    
    # Get full expertise text
    full_text = expertise_info.get("full_content", "")
    
    # Extract themes and capabilities as separate topics
    themes = expertise_info.get("themes", [])
    capabilities = expertise_info.get("capabilities", [])
    
    # If we have multiple themes/capabilities, treat each as a topic
    if len(themes) > 1 or len(capabilities) > 1:
        for theme in themes:
            # Find text around this theme in the description
            topic_text = theme + " " + " ".join([cap for cap in capabilities if any(word in cap.lower() for word in theme.lower().split())])
            topics.append({
                "text": topic_text,
                "label": theme[:50]
            })
    else:
        # Single research line - treat entire expertise as one topic
        topics.append({
            "text": full_text,
            "label": "Primary Expertise"
        })
    
    return topics

def check_domain_keyword_filter(expertise_info, call_text, threshold_keywords=5, threshold_overlap=0.10):
    """
    PASS 1: Hard keyword filter to eliminate generic term false positives.
    
    Extracts domain keywords from expertise and checks overlap with call description.
    Uses intelligent keyword extraction:
    - Extracts meaningful multi-word phrases (self-healing, soft robotics)
    - Excludes very generic technical terms (systems, development, advanced)
    - Weights keyword matches by domain specificity
    
    Rejects if:
    - Less than threshold_keywords (default 5) SPECIFIC keywords present, AND
    - Less than threshold_overlap (default 10%) of expertise keywords present
    
    Returns dict with:
    - passes: bool (True if passes keyword filter)
    - matching_keywords: list of keywords found in call
    - expertise_keywords: set of domain keywords from expertise
    - overlap_percentage: % of expertise keywords found in call
    """
    
    # Extract domain keywords from expertise with focus on specificity
    expertise_keywords = set()
    generic_terms = {
        'and', 'the', 'for', 'with', 'from', 'that', 'this', 'are', 'was', 'were', 'been', 'have', 'has', 'can', 'will',
        'system', 'systems', 'development', 'advanced', 'method', 'methods', 'approach', 'approaches',
        'solution', 'solutions', 'framework', 'frameworks', 'technology', 'technologies', 'platform', 'platforms',
        'performance', 'application', 'applications', 'based', 'enable', 'enables', 'provide', 'provides'
    }
    
    # Extract from themes - prefer multi-word phrases
    for theme in expertise_info.get("themes", []):
        # Split on common separators
        words = re.split(r'[\s,\-&()]', theme.lower())
        words = [w.strip() for w in words if w.strip() and len(w) >= 3]
        
        # Group into phrase terms when possible (2-3 word combinations)
        phrase_parts = [w for w in words if w not in generic_terms]
        
        # Add longer meaningful phrases (2+ words)
        for i in range(len(phrase_parts) - 1):
            if len(phrase_parts[i]) >= 4 and len(phrase_parts[i + 1]) >= 4:
                expertise_keywords.add(f"{phrase_parts[i]} {phrase_parts[i + 1]}")
        
        # Add individual significant words (5+ chars to be more selective)
        for word in phrase_parts:
            if len(word) >= 5:  # More selective: longer words are more specific
                expertise_keywords.add(word)
    
    # Extract from capabilities with same logic
    for cap in expertise_info.get("capabilities", []):
        words = re.split(r'[\s,\-&()]', cap.lower())
        words = [w.strip() for w in words if w.strip() and len(w) >= 3]
        
        phrase_parts = [w for w in words if w not in generic_terms]
        
        # Add longer phrases
        for i in range(len(phrase_parts) - 1):
            if len(phrase_parts[i]) >= 4 and len(phrase_parts[i + 1]) >= 4:
                expertise_keywords.add(f"{phrase_parts[i]} {phrase_parts[i + 1]}")
        
        # Add individual significant words (5+ chars)
        for word in phrase_parts:
            if len(word) >= 5:
                expertise_keywords.add(word)
    
    # Extract keywords from call - also emphasizing longer, more specific terms
    call_words = set()
    for word in re.findall(r'\b[a-z]{3,}\b', call_text.lower()):
        if word not in generic_terms and len(word) >= 4:  # Only specific terms
            call_words.add(word)
    
    # Also extract multi-word phrases from call
    call_phrases = set()
    call_word_tokens = call_text.lower().split()
    for i in range(len(call_word_tokens) - 1):
        w1 = call_word_tokens[i].lower().strip('.,;:')
        w2 = call_word_tokens[i + 1].lower().strip('.,;:')
        if w1 not in generic_terms and w2 not in generic_terms and len(w1) >= 4 and len(w2) >= 4:
            call_phrases.add(f"{w1} {w2}")
    
    call_words.update(call_phrases)
    
    # Count matching keywords
    matching_keywords = expertise_keywords.intersection(call_words)
    
    # Calculate overlap percentage
    overlap_percentage = len(matching_keywords) / max(1, len(expertise_keywords)) if expertise_keywords else 0.0
    
    # Determine if passes filter - BOTH conditions must be met for domain relevance
    passes = len(matching_keywords) >= threshold_keywords or overlap_percentage >= threshold_overlap
    
    return {
        "passes": passes,
        "matching_keywords": sorted(list(matching_keywords)),
        "expertise_keywords": expertise_keywords,
        "expertise_keyword_count": len(expertise_keywords),
        "overlap_percentage": overlap_percentage,
        "threshold_keywords": threshold_keywords,
        "threshold_overlap": threshold_overlap
    }

def extract_technical_keywords_from_call(call_text):
    """
    Extract technical keywords and phrases from call description.
    Returns set of meaningful technical terms that should appear in expertise.
    
    Focuses on:
    - Technical nouns and noun phrases (2-4 words)
    - Domain-specific terminology
    - Action-oriented technical terms
    """
    if not call_text or len(call_text) < 20:
        return set()
    
    # Convert to lowercase for matching
    text = call_text.lower()
    
    # Extract multi-word technical phrases (2-4 words)
    # Common patterns: adjective + noun, noun + noun, etc.
    technical_phrases = []
    
    # Pattern 1: Hyphenated technical terms
    hyphenated = re.findall(r'\b[a-z]+(?:-[a-z]+)+\b', text)
    technical_phrases.extend(hyphenated)
    
    # Pattern 2: Multi-word technical terms (using common technical patterns)
    bigrams = re.findall(r'\b(?:artificial intelligence|machine learning|deep learning|computer vision|'
                        r'natural language|human-robot|collaborative robot|autonomous navigation|'
                        r'digital twin|smart manufacturing|advanced materials|renewable energy|'
                        r'supply chain|climate change|data analytics|sensor fusion|motion planning|'
                        r'reinforcement learning|neural network|predictive maintenance|quality control|'
                        r'additive manufacturing|circular economy|sustainability|robotics|automation|'
                        r'internet of things|cyber security|edge computing|cloud computing|'
                        r'virtual reality|augmented reality|mixed reality|wearable|exoskeleton|'
                        r'rehabilitation|assistive technology|mobility|gait analysis|biomechanics|'
                        r'healthcare|medical device|surgical robot|diagnostic|therapeutic|'
                        r'industrial robot|service robot|mobile robot|manipulator|gripper|actuator|'
                        r'control system|embedded system|real-time|sensor network|wireless|'
                        r'battery|energy storage|power management|charging|electric vehicle|'
                        r'biofuel|hydrogen|solar|wind|photovoltaic|thermal|'
                        r'material science|composite|polymer|alloy|nano|coating|'
                        r'manufacturing|production|assembly|logistics|warehouse|transportation|'
                        r'optimization|simulation|modeling|validation|testing|certification)\b', text)
    technical_phrases.extend(bigrams)
    
    # Pattern 3: Technical acronyms (2-6 uppercase letters)
    acronyms = re.findall(r'\b[A-Z]{2,6}\b', call_text)  # Use original case for acronyms
    technical_phrases.extend([a.lower() for a in acronyms if a not in ['EU', 'USA', 'UK', 'EUR', 'USA']])
    
    # Extract single significant technical words (4+ chars, not common words)
    words = re.findall(r'\b[a-z]{4,}\b', text)
    
    # Filter out common non-technical words
    stop_words = {
        'that', 'this', 'with', 'from', 'have', 'will', 'been', 'were', 'their', 'about',
        'would', 'there', 'which', 'other', 'these', 'should', 'could', 'also', 'more',
        'such', 'when', 'than', 'some', 'into', 'through', 'between', 'under', 'over',
        'those', 'while', 'where', 'after', 'before', 'during', 'without', 'within',
        'action', 'research', 'innovation', 'project', 'topic', 'proposal', 'expected',
        'development', 'support', 'framework', 'approach', 'focus', 'objective', 'scope',
        'include', 'provide', 'ensure', 'demonstrate', 'address', 'enable', 'develop'
    }
    
    technical_words = [w for w in words if w not in stop_words and w not in ['and', 'the', 'for', 'are', 'was']]
    
    # Combine all keywords
    all_keywords = set(technical_phrases + technical_words)
    
    # Keep only meaningful keywords (at least 3 chars, not purely common)
    filtered_keywords = {kw for kw in all_keywords if len(kw) >= 3}
    
    return filtered_keywords

def extract_and_score_projects(expertise_info, call_text):
    """
    Extract individual projects and score each against the call.
    Returns BEST matching project score (not average).
    
    Key insight: 1 strong match > 10 weak matches.
    """
    # Extract "Previous Projects" or "Previous Involved Projects" section
    # Support multiple formats from different expertise file structures
    projects_section = ""
    full_text = expertise_info.get("full_content", "")
    
    # Try multiple patterns for project sections
    patterns = [
        r"##?\s*PREVIOUS PROJECTS[:\s]*([\s\S]*?)(?=\n##|\nPROFESSOR:|$)",  # Template format
        r"Previous Involved Projects[:\s]*([\s\S]*?)(?=\n\n[A-Z]|Keywords|$)",  # Old format
        r"Previous Projects[:\s]*([\s\S]*?)(?=\n\n[A-Z]|Keywords|$)",  # Variant
        r"Projects[:\s]*\n([\s\S]*?)(?=\n\n[A-Z#]|Keywords|$)",  # Simple "Projects:"
    ]
    
    for pattern in patterns:
        match = re.search(pattern, full_text, re.IGNORECASE)
        if match:
            projects_section = match.group(1)
            break
    
    if not projects_section:
        # Fallback: No projects found at all
        return {"best_score": 0, "best_project": None, "count": 0}
    
    # Split into individual projects (assume each starts with "###", "- ", "• " or number)
    project_lines = []
    for line in projects_section.split("\n"):
        line = line.strip()
        # Match project headers: "###", "- ", "• ", numbered lists, or "Project N:"
        if line and (line.startswith("###") or line.startswith("-") or line.startswith("•") or 
                    (line and line[0].isdigit() and "." in line[:3]) or 
                    re.match(r"^Project \d+:", line, re.IGNORECASE)):
            project_lines.append(line)
    
    if not project_lines:
        # Fallback: treat entire section as one project
        project_lines = [projects_section]
    
    # Score each project individually
    project_scores = []
    for project in project_lines:
        score = compute_semantic_similarity(project.lower(), call_text) * 100
        project_scores.append({"text": project[:100], "score": score})
    
    # Return BEST match
    if project_scores:
        best = max(project_scores, key=lambda x: x["score"])
        return {
            "best_score": best["score"],
            "best_project": best["text"],
            "count": len(project_scores)
        }
    
    return {"best_score": 0, "best_project": None, "count": 0}

def calculate_continuous_relevance_score(call_title, call_description, expertise_info):
    """
    3-PASS SCORING SYSTEM with keyword filtering:
    
    PASS 1: Domain Keyword Filter
    - Extract expertise keywords from themes/capabilities
    - Check overlap with call text
    - REJECT if <5 keywords or <10% overlap (hard veto)
    - Eliminates false positives on generic technical vocabulary
    
    PASS 2: Semantic Similarity (only if keyword filter passed)
    - Use embedding-based similarity (sentence-transformers)
    - Require >60% similarity (higher threshold for domain-qualified calls)
    - Score best matching topic independently
    
    PASS 3: Enhanced Bidirectional Keyword Matching + Adaptive Project Weighting
    - Extract technical keywords FROM CALL DESCRIPTION
    - Check if they appear in FULL EXPERTISE CONTENT (not just themes)
    - Select best topic score (not average)
    - Select best project score (not average)
    - Adaptive weighting based on project match quality (realistic thresholds):
      * Strong project (≥40%): 25% semantic + 25% domain + 50% project
      * Good project (≥30%): 35% semantic + 30% domain + 35% project
      * Weak project (<30%): 50% semantic + 35% domain + 15% project
      * No projects: 50% semantic + 35% domain
    
    Returns dict with score and components.
    """
    call_text = f"{call_title} {call_description}".lower()
    expertise_text = expertise_info.get("full_content", "").lower()
    
    # PASS 1: Domain Keyword Filter
    keyword_filter_result = check_domain_keyword_filter(expertise_info, call_text)
    
    # If keyword filter fails, return low score
    if not keyword_filter_result["passes"]:
        return {
            "score": 0,  # Hard reject
            "semantic": 0,
            "domain": 0,
            "project": 0,
            "best_topic": "Rejected by keyword filter",
            "best_project": None,
            "match_rationale": "keyword_filter_fail",
            "topic_count": 0,
            "project_count": 0,
            "components": {"keyword_filter": 0.0},
            "keyword_filter": keyword_filter_result
        }
    
    # PASS 2: Topic-based semantic scoring (find BEST matching topic)
    topics = segment_expertise_into_topics(expertise_info)
    topic_scores = []
    
    for topic in topics:
        score = compute_semantic_similarity(topic["text"].lower(), call_text) * 100
        topic_scores.append({"label": topic["label"], "score": score})
    
    # Take BEST topic score (not average)
    best_topic = max(topic_scores, key=lambda x: x["score"]) if topic_scores else {"score": 0, "label": "None"}
    semantic_score = best_topic["score"]
    
    # PASS 2b: Apply higher semantic threshold for domain-qualified calls
    # Only require >60% if keyword filter passed (domain keywords present)
    if semantic_score < 60:
        # Weak semantic match even after keyword filter passes
        # This is allowable if project evidence is strong
        pass
    
    # Step 3: ENHANCED Domain matching (bidirectional keyword matching)
    # Extract technical keywords from CALL DESCRIPTION (not just from expertise)
    # Then check if they appear in FULL EXPERTISE CONTENT
    # This leverages rich description content to find relevant experts
    
    call_keywords = extract_technical_keywords_from_call(call_text)
    
    if not call_keywords:
        domain_score = 0
    else:
        # Check how many call keywords appear in expertise full content
        expertise_text = expertise_info.get("full_content", "").lower()
        
        # Count matches with slight flexibility (allow partial word matches for compounds)
        matched_keywords = set()
        for keyword in call_keywords:
            if keyword in expertise_text:
                matched_keywords.add(keyword)
            # Also check for partial matches in compound words (e.g., "robot" matches "robotics")
            elif len(keyword) >= 5 and any(keyword in word for word in expertise_text.split()):
                matched_keywords.add(keyword)
        
        # Score based on coverage of call keywords
        coverage = len(matched_keywords) / len(call_keywords)
        
        # Apply non-linear scaling: reward high coverage more
        # 50% coverage = 70 score, 70% coverage = 85 score, 90% coverage = 95 score
        domain_score = min(100, coverage * 120)  # Scale up to emphasize importance
    
    # Step 3: Project matching (find BEST matching project)
    project_result = extract_and_score_projects(expertise_info, call_text)
    project_score = project_result["best_score"]
    
    # Step 4: ADAPTIVE WEIGHTING based on project strength
    # Four scenarios prioritizing strong project evidence (realistic thresholds):
    #   A. Strong project match (≥40%): Major project weight - best predictor
    #   B. Good project match (≥30%): High project weight - strong signal
    #   C. Weak project match (<30%): Low project weight - current expertise focus
    #   D. No projects: Pure current expertise pathway
    
    if project_result["count"] == 0:
        # No projects found: do not penalize missing project history.
        # Use full topic similarity score as requested.
        final_score = semantic_score
        match_rationale = "current_expertise_only"
        weights_used = {"semantic": 1.00, "project": 0.00, "domain": 0.00}
    elif project_score >= 40:
        # Path B: Strong project match (realistically >=40% based on actual data)
        # Strong project evidence gets major weight!
        # Weights: 25% semantic + 25% domain + 50% best_project
        final_score = (semantic_score * 0.25) + (domain_score * 0.25) + (project_score * 0.50)
        match_rationale = "strong_project_precedent"
        weights_used = {"semantic": 0.25, "project": 0.50, "domain": 0.25}
    elif project_score >= 30:
        # Path C: Moderate project match (>=30% based on actual data distribution)
        # Meaningful project evidence provides good support
        # Weights: 35% semantic + 30% domain + 35% best_project
        final_score = (semantic_score * 0.35) + (domain_score * 0.30) + (project_score * 0.35)
        match_rationale = "good_project_match"
        weights_used = {"semantic": 0.35, "project": 0.35, "domain": 0.30}
    else:
        # Path A: Weak project match (<30%)
        # Weights: 50% semantic + 35% domain + 15% best_project
        final_score = (semantic_score * 0.50) + (domain_score * 0.35) + (project_score * 0.15)
        match_rationale = "current_expertise"
        weights_used = {"semantic": 0.50, "project": 0.15, "domain": 0.35}
    
    # Clamp to 0-100
    final_score = max(0, min(100, final_score))
    
    return {
        "score": final_score,  # 0-100
        "semantic": semantic_score,  # 0-100
        "domain": domain_score,  # 0-100
        "project": project_score,  # 0-100
        "best_topic": best_topic["label"],
        "best_project": project_result["best_project"],
        "match_rationale": match_rationale,
        "topic_count": len(topics),
        "project_count": project_result["count"],
        "components": weights_used,
        "keyword_filter": keyword_filter_result  # Include keyword filter details
    }

def extract_expertise_description(expertise_file):
    """
    Extract expertise core description, capabilities, and themes from any text file.
    Focuses on semantic meaning, not just keywords.
    Enhanced with n-gram extraction and technical terminology detection.
    """
    with open(expertise_file, 'r', encoding='utf-8') as f:
        content = f.read()
    
    expertise_name = expertise_file.stem
    
    # Extract key sections and capabilities
    lines = content.split('\n')
    
    # Collect all non-empty lines for analysis
    meaningful_lines = [line.strip() for line in lines if line.strip() and len(line.strip()) > 10]
    
    # Extract themes by analyzing sentence structure
    themes = []
    capabilities = []
    technical_terms = []
    
    for line in meaningful_lines:
        line_lower = line.lower()
        
        # Detect capability statements
        if any(word in line_lower for word in ["develop", "enable", "provide", "create", "implement", "design", "improve", "enhance", "integrate", "achieve", "deliver"]):
            capabilities.append(line)
        
        # Detect theme statements
        if any(word in line_lower for word in ["focus", "aim", "goal", "mission", "purpose", "specializ", "expert", "core", "strength"]):
            themes.append(line)
        
        # Extract technical terms (multi-word phrases with specific patterns)
        tech_patterns = re.findall(r'\b(?:[A-Z][a-z]+\s){1,3}(?:system|technology|platform|solution|framework|algorithm|method|approach)\b', line)
        technical_terms.extend(tech_patterns)
    
    # If not enough explicit sections, use full content for qualitative analysis
    if not themes:
        themes = meaningful_lines[:3]
    if not capabilities:
        capabilities = meaningful_lines[3:6] if len(meaningful_lines) > 3 else meaningful_lines
    
    # Extract bigrams and trigrams for better phrase matching
    content_lower = content.lower()
    words = re.findall(r'\b\w{4,}\b', content_lower)
    bigrams = [f"{words[i]} {words[i+1]}" for i in range(len(words)-1)] if len(words) > 1 else []
    trigrams = [f"{words[i]} {words[i+1]} {words[i+2]}" for i in range(len(words)-2)] if len(words) > 2 else []
    
    return {
        "name": expertise_name,
        "full_content": content,
        "themes": themes,
        "capabilities": capabilities,
        "technical_terms": technical_terms,
        "content_length": len(content),
        "content_lower": content_lower,
        "bigrams": set(bigrams),
        "trigrams": set(trigrams),
        "words": set(words)
    }

def validate_alignment_pass1_semantic_similarity(call_title, call_description, expertise_info):
    """
    First validation pass: Semantic similarity using fuzzy token matching.
    Replaces word-overlap approach with token-based matching that handles:
    - Cross-domain expertise (e.g., haptics in robotics vs VR)
    - Vocabulary differences between domains
    - Partial matching (substrings, core concepts)
    
    Scoring Strategy:
    - 70% weight on objective/main description similarity
    - 30% weight on keyword similarity (with confidence based on keyword count/diversity)
    """
    call_text = f"{call_title} {call_description}"
    call_description_only = call_description
    
    # Extract expertise description/objectives (use themes + capabilities)
    expertise_themes_text = " ".join(expertise_info.get("themes", []))
    expertise_capabilities_text = " ".join(expertise_info.get("capabilities", []))
    expertise_description = expertise_themes_text + " " + expertise_capabilities_text
    
    if not expertise_description.strip():
        expertise_description = expertise_info.get("full_content", "")
    
    # 1. OBJECTIVE SIMILARITY (70% weight)
    # Compare main call description with expertise description using semantic matching
    objective_similarity = compute_semantic_similarity(call_description_only, expertise_description)
    
    # 2. KEYWORD SIMILARITY (30% weight)
    # Extract keywords from full content if available (look for "Keywords:" section)
    expertise_keywords = []
    full_content = expertise_info.get("full_content", "").lower()
    
    # Try to find keywords section in expertise
    kw_match = re.search(r"(?:keywords|key terms)[:\s]*(.*?)(?:\n\n|\Z)", full_content, re.IGNORECASE | re.DOTALL)
    if kw_match:
        kw_text = kw_match.group(1)
        # Split by comma or newline, clean up
        expertise_keywords = [k.strip() for k in re.split(r'[,\n]', kw_text) if k.strip() and len(k.strip()) > 2]
    
    # Try to find keywords in call description
    call_keywords = []
    call_text_lower = call_text.lower()
    call_kw_match = re.search(r"(?:keywords|key terms)[:\s]*(.*?)(?:\n\n|[.!?]|\Z)", call_text_lower, re.IGNORECASE | re.DOTALL)
    if call_kw_match:
        kw_text = call_kw_match.group(1)
        call_keywords = [k.strip() for k in re.split(r'[,\n]', kw_text) if k.strip() and len(k.strip()) > 2]
    
    keyword_similarity = 0.0
    keyword_confidence = 0.0
    
    if expertise_keywords and call_keywords:
        # Use semantic matching on keywords
        expertise_kw_text = " ".join(expertise_keywords)
        call_kw_text = " ".join(call_keywords)
        keyword_similarity = compute_semantic_similarity(call_kw_text, expertise_kw_text)
        
        # Keyword confidence: based on count, length, and diversity
        num_keywords = len(expertise_keywords)
        avg_len = sum(len(k) for k in expertise_keywords) / num_keywords if num_keywords else 0
        diversity = len(set(expertise_keywords)) / (num_keywords or 1)
        
        # Confidence formula: 0.5 * (keywords/10) + 0.3 * (avg_len/10) + 0.2 * diversity
        keyword_confidence = min(1.0, 0.5 * (num_keywords / 10) + 0.3 * (avg_len / 10) + 0.2 * diversity)
    
    # 3. COMBINED SCORE
    # Use 70% objective + 30% keywords (weighted by keyword confidence)
    w_objective = 0.7
    w_keywords = 0.3 * keyword_confidence
    total_weight = w_objective + w_keywords
    
    if total_weight > 0:
        final_score = (w_objective * objective_similarity + w_keywords * keyword_similarity) / total_weight
    else:
        final_score = objective_similarity
    
    # 4. DETERMINE PASS (threshold: 0.30 for fuzzy token matching)
    # This is more forgiving than word overlap but more strict than pure embeddings
    passes = final_score >= 0.30
    
    return {
        "pass": passes,
        "score": final_score,
        "objective_similarity": objective_similarity,
        "keyword_similarity": keyword_similarity,
        "keyword_confidence": keyword_confidence,
        "method": "fuzzy-token-based"
    }

def validate_alignment_pass2_thematic_match(call_title, call_description, expertise_info):
    """
    Second validation pass: Enhanced thematic alignment with TRL awareness and domain clustering.
    Identifies not just domain presence but strength and contextual relevance.
    """
    call_text = f"{call_title} {call_description}".lower()
    
    # Expanded thematic domains with more comprehensive keywords
    thematic_domains = {
        "manufacturing": ["manufacturing", "production", "factory", "industrial", "assembly", "automation", "process", "fabrication", "shop floor", "production line"],
        "robotics": ["robot", "robotic", "collaborative", "cobot", "autonomous", "manipulator", "mobile robot", "arm", "gripper", "end effector"],
        "safety": ["safety", "safe", "protection", "risk", "secure", "wellbeing", "hazard", "human-robot", "interaction", "collision", "proximity"],
        "human_focus": ["worker", "operator", "human", "workforce", "people", "staff", "employee", "ergonomic", "physical", "cognitive load", "augmentation"],
        "innovation": ["innovation", "novel", "breakthrough", "advanced", "emerging", "disruptive", "state-of-the-art", "cutting-edge"],
        "training": ["training", "skill", "learning", "knowledge", "expertise", "development", "apprenticeship", "upskilling", "education"],
        "ai_ml": ["artificial intelligence", "machine learning", "deep learning", "neural", "algorithm", "data-driven", "prediction", "classification"],
        "technology": ["technology", "digital", "intelligent", "smart", "iot", "cyber-physical", "digitalization", "industry 4.0"],
        "sustainability": ["sustainable", "circular", "green", "environmental", "reuse", "recycling", "lifecycle", "carbon", "eco"],
        "materials": ["material", "polymer", "composite", "additive", "healing", "recyclable", "alloy", "metal", "self-healing", "biodegradable"],
        "hardware": ["hardware", "wearable", "sensor", "actuator", "haptic", "device", "equipment", "interface", "headset", "component", "embedded"],
        "ar_vr": ["augmented reality", "virtual reality", "mixed reality", "immersive", "visualization", "simulation", "ar", "vr", "xr"],
        "data_analytics": ["data", "analytics", "monitoring", "tracking", "measurement", "diagnostic", "predictive maintenance"],
        "networking": ["network", "communication", "connectivity", "wireless", "protocol", "5g", "6g", "data transfer"],
        "energy": ["energy", "power", "battery", "consumption", "efficiency", "renewable", "storage"]
    }
    
    # Domain importance weights (specific technical domains weighted higher)
    domain_weights = {
        "manufacturing": 1.3,
        "robotics": 1.5,
        "safety": 1.3,
        "human_focus": 1.2,
        "innovation": 0.8,
        "training": 1.1,
        "ai_ml": 1.4,
        "technology": 0.8,
        "sustainability": 1.1,
        "materials": 1.4,
        "hardware": 1.4,
        "ar_vr": 1.3,
        "data_analytics": 1.2,
        "networking": 1.1,
        "energy": 1.0
    }
    
    # TRL indicators (technology readiness level context)
    trl_indicators = {
        "high": ["deployment", "commercial", "market", "pilot", "demonstration", "validation", "trl 7", "trl 8", "trl 9"],
        "medium": ["prototype", "testing", "integration", "trl 5", "trl 6", "development"],
        "low": ["research", "concept", "feasibility", "trl 3", "trl 4", "proof of concept"]
    }
    
    # Check which domains are present in call with strength scoring
    call_domain_strength = {}
    for domain, keywords in thematic_domains.items():
        domain_matches = sum(1 for kw in keywords if kw in call_text)
        # Strength: proportion of keywords found, with bonus for multiple occurrences
        total_occurrences = sum(call_text.count(kw) for kw in keywords if kw in call_text)
        strength = min((domain_matches / len(keywords)) + (total_occurrences * 0.01), 1.0)
        call_domain_strength[domain] = strength
    
    # Check which domains are present in expertise
    expertise_domain_strength = {}
    expertise_text = expertise_info["content_lower"]
    for domain, keywords in thematic_domains.items():
        domain_matches = sum(1 for kw in keywords if kw in expertise_text)
        total_occurrences = sum(expertise_text.count(kw) for kw in keywords if kw in expertise_text)
        strength = min((domain_matches / len(keywords)) + (total_occurrences * 0.01), 1.0)
        expertise_domain_strength[domain] = strength
    
    # Detect TRL alignment
    call_trl = "unknown"
    for trl_level, indicators in trl_indicators.items():
        if any(ind in call_text for ind in indicators):
            call_trl = trl_level
            break
    
    expertise_trl = "unknown"
    for trl_level, indicators in trl_indicators.items():
        if any(ind in expertise_text for ind in indicators):
            expertise_trl = trl_level
            break
    
    # TRL alignment bonus (if both are at similar readiness levels)
    trl_bonus = 0
    if call_trl != "unknown" and expertise_trl != "unknown":
        if call_trl == expertise_trl:
            trl_bonus = 0.1  # 10% bonus for matching TRL
        elif abs(["low", "medium", "high"].index(call_trl) - ["low", "medium", "high"].index(expertise_trl)) == 1:
            trl_bonus = 0.05  # 5% bonus for adjacent TRL
    
    # Calculate weighted domain overlap
    weighted_overlap = 0
    shared_domains_with_strength = []
    
    for domain in thematic_domains:
        call_strength = call_domain_strength[domain]
        expertise_strength = expertise_domain_strength[domain]
        
        # Both must be present AND both must be significant (>20% of keywords)
        if call_strength > 0.2 and expertise_strength > 0.2:
            # Weight by importance and geometric mean of strengths
            overlap = (call_strength * expertise_strength) ** 0.5 * domain_weights[domain]
            weighted_overlap += overlap
            shared_domains_with_strength.append((domain, call_strength, expertise_strength))
    
    # Normalize
    max_possible_overlap = sum(domain_weights.values()) * 1.0
    thematic_alignment = (weighted_overlap / max_possible_overlap) + trl_bonus
    
    # Count strong shared domains (both > 30%)
    num_strong_shared = sum(1 for domain, cs, es in shared_domains_with_strength if cs > 0.3 and es > 0.3)
    
    return {
        "pass": weighted_overlap >= 0.8 and len(shared_domains_with_strength) >= 1,  # Balanced: lower overlap threshold, 1+ shared domain
        "score": min(thematic_alignment, 1.0),
        "domain_overlap": len(shared_domains_with_strength),
        "strong_domains": num_strong_shared,
        "weighted_overlap": weighted_overlap,
        "trl_match": call_trl == expertise_trl if call_trl != "unknown" and expertise_trl != "unknown" else None,
        "call_domains": call_domain_strength,
        "shared_domains": [d for d, cs, es in shared_domains_with_strength]
    }

def validate_alignment_pass3_content_relevance(call_title, call_description, expertise_info):
    """
    Third validation pass: Enhanced capability-problem matching with contextual understanding.
    Validates that expertise substantively addresses call requirements with actionable solutions.
    """
    call_text = f"{call_title} {call_description}".lower()
    
    # Extract capability themes from expertise
    expertise_full = expertise_info["full_content"].lower()
    themes_text = " ".join(expertise_info["themes"]).lower()
    capabilities_text = " ".join(expertise_info["capabilities"]).lower()
    combined_expertise = f"{themes_text} {capabilities_text}"
    
    # Enhanced problem indicators with contextual categories
    problem_indicators = {
        "challenges": ["challenge", "problem", "difficulty", "complex", "complicated", "obstacle", "barrier"],
        "needs": ["need", "require", "requirement", "lack", "lacking", "absence", "missing", "essential"],
        "goals": ["improve", "enhance", "optimize", "strengthen", "advance", "increase", "maximize", "achieve"],
        "gaps": ["gap", "shortage", "limitation", "constraint", "barrier", "bottleneck", "insufficient"],
        "demands": ["demand", "expect", "must", "should", "shall", "critical", "important", "necessary"]
    }
    
    # Enhanced solution indicators with action verbs
    solution_indicators = {
        "develops": ["develop", "create", "design", "build", "construct", "engineer", "architect"],
        "provides": ["provide", "offer", "deliver", "supply", "enable", "facilitate", "support"],
        "implements": ["implement", "apply", "deploy", "integrate", "install", "execute", "operationalize"],
        "improves": ["improve", "enhance", "optimize", "strengthen", "advance", "upgrade", "refine"],
        "validates": ["validate", "verify", "test", "demonstrate", "prove", "assess", "evaluate"]
    }
    
    # Count problem indicators in call (with context awareness)
    problem_count = 0
    problem_types_found = []
    problem_contexts = []
    
    for ptype, keywords in problem_indicators.items():
        for kw in keywords:
            if kw in call_text:
                # Extract context around the keyword (50 chars before and after)
                idx = call_text.find(kw)
                if idx != -1:
                    context = call_text[max(0, idx-50):min(len(call_text), idx+50)]
                    problem_contexts.append(context)
                    problem_count += 1
        if any(kw in call_text for kw in keywords):
            problem_types_found.append(ptype)
    
    # Count solution indicators in expertise
    solution_count = 0
    solution_types_found = []
    solution_contexts = []
    
    for stype, keywords in solution_indicators.items():
        for kw in keywords:
            if kw in combined_expertise:
                idx = combined_expertise.find(kw)
                if idx != -1:
                    context = combined_expertise[max(0, idx-50):min(len(combined_expertise), idx+50)]
                    solution_contexts.append(context)
                    solution_count += 1
        if any(kw in combined_expertise for kw in keywords):
            solution_types_found.append(stype)
    
    # Deep capability matching - check if expertise capabilities address call needs
    capability_relevance = 0
    direct_matches = 0
    
    # Extract sentences from call describing needs/goals
    call_sentences = call_text.replace('.', '\n').replace('!', '\n').replace('?', '\n').split('\n')
    need_sentences = [s.strip() for s in call_sentences if len(s.strip()) > 20 and any(p in s for prob_list in problem_indicators.values() for p in prob_list)]
    
    # Check if expertise capabilities address extracted needs
    for need in need_sentences:
        need_words = set(w for w in re.findall(r'\b\w{4,}\b', need) if len(w) > 3)
        expertise_words = set(re.findall(r'\b\w{4,}\b', combined_expertise))
        
        overlap = len(need_words & expertise_words)
        if overlap > 3:  # Significant overlap
            capability_relevance += 1
            if overlap > 5:  # Strong overlap
                direct_matches += 1
    
    # Contextual alignment: Check if problem contexts and solution contexts have semantic overlap
    context_alignment = 0
    for prob_context in problem_contexts[:5]:  # Check top 5 problem contexts
        prob_words = set(re.findall(r'\b\w{4,}\b', prob_context))
        for sol_context in solution_contexts[:5]:  # Against top 5 solution contexts
            sol_words = set(re.findall(r'\b\w{4,}\b', sol_context))
            if len(prob_words & sol_words) > 2:  # Contextual overlap
                context_alignment += 1
    
    # Scoring components
    problem_score = min(problem_count / 4, 1.0) if problem_count > 0 else 0
    solution_score = min(solution_count / 3, 1.0) if solution_count > 0 else 0
    capability_score = min(capability_relevance / 3, 1.0) if capability_relevance > 0 else 0
    context_score = min(context_alignment / 5, 1.0) if context_alignment > 0 else 0
    direct_match_bonus = min(direct_matches * 0.1, 0.2)  # Up to 20% bonus for direct matches
    
    # Combined relevance score with emphasis on actionable solutions
    combined_relevance = (
        problem_score * 0.20 +       # Problems identified
        solution_score * 0.30 +       # Solutions available
        capability_score * 0.25 +     # Capability alignment
        context_score * 0.15 +        # Contextual alignment
        direct_match_bonus * 0.10     # Direct match bonus
    )
    
    # Pass requires: problems, solutions, capability relevance (balanced thresholds)
    has_problems = problem_count >= 2  # At least 2 problem indicators
    has_solutions = solution_count >= 2  # At least 2 solution indicators  
    has_capabilities = capability_relevance >= 1  # At least 1 capability match
    strong_relevance = combined_relevance >= 0.30  # Balanced threshold
    
    return {
        "pass": has_problems and has_solutions and has_capabilities and strong_relevance,
        "score": combined_relevance,
        "problem_count": problem_count,
        "solution_count": solution_count,
        "capability_relevance": capability_relevance,
        "direct_matches": direct_matches,
        "context_alignment": context_alignment,
        "problem_score": problem_score,
        "solution_score": solution_score
    }

def validate_alignment_comprehensive(call_title, call_description, expertise_info):
    """
    SIMPLIFIED APPROACH: Calculate continuous relevance score (0-100).
    
    No more complex multi-pass validation with thresholds.
    Just: semantic (70%) + domains (15%) + projects (15%) = score
    
    Returns score for ranking and selection.
    """
    result = calculate_continuous_relevance_score(call_title, call_description, expertise_info)
    score = result["score"] / 100.0  # Convert back to 0-1 for compatibility
    
    # Threshold raised to 35 to reduce noise
    overall_pass = result["score"] > 35.0
    
    # Quality classification based on score
    if result["score"] >= 70:
        match_quality = "excellent"
    elif result["score"] >= 55:
        match_quality = "good"
    elif result["score"] >= 40:
        match_quality = "fair"
    else:
        match_quality = "none"
    
    return {
        "overall_pass": overall_pass,
        "confidence": score,  # Keep as 0-1 for compatibility
        "match_quality": match_quality,
        "score_0_100": result["score"],  # NEW: Score out of 100
        "passes_count": 1,  # Simplified - single continuous score
        "semantic_similarity": {
            "pass": result["semantic"] > 25,
            "score": result["semantic"] / 100.0
        },
        "thematic_match": {"pass": True, "score": 0.0},  # Deprecated
        "content_relevance": {"pass": True, "score": 0.0},  # Deprecated
        "shared_domains": []
    }

def generate_qualitative_explanation(expertise_info, validation_results, call_title, call_description):

    # Identify top contributors to the score
    semantic = validation_results.get("semantic_similarity", {})
    thematic = validation_results.get("thematic_match", {})
    content = validation_results.get("content_relevance", {})

    contributors = []
    if semantic.get("score", 0) > 0.18:
        contributors.append(f"semantic similarity ({semantic.get('score', 0):.0%})")
    if thematic.get("score", 0) > 0.18 and thematic.get("domain_overlap", 0) > 0:
        doms = ', '.join(thematic.get("shared_domains", [])[:2])
        contributors.append(f"thematic overlap in {doms} ({thematic.get('score', 0):.0%})")
    if content.get("score", 0) > 0.18:
        contributors.append(f"problem-solution relevance ({content.get('score', 0):.0%})")

    """
    Generate a natural-language project alignment explanation by extracting and combining full sentences from the expertise and call, similar to the call description fetcher.
    """
    shared_domains = validation_results.get("shared_domains", [])
    confidence = validation_results.get("confidence", 0)
    match_quality = validation_results.get("match_quality", "unknown")
    content_relevance = validation_results.get("content_relevance", {})
    thematic_match = validation_results.get("thematic_match", {})
    expertise_name = expertise_info['name']
    expertise_themes = expertise_info.get('themes', [])
    expertise_capabilities = expertise_info.get('capabilities', [])
    call_excerpt = call_description[:600]

    # Extract 1-2 meaningful sentences from expertise capabilities
    cap_sentences = []
    for cap in expertise_capabilities:
        # Use only sentences with at least 8 words and not too generic
        for s in re.split(r'(?<=[.!?])\s+', cap):
            if len(s.split()) >= 8 and not s.lower().startswith("we "):
                cap_sentences.append(s.strip())
            if len(cap_sentences) >= 2:
                break
        if len(cap_sentences) >= 2:
            break

    # Extract 1-2 meaningful sentences from expertise themes
    theme_sentences = []
    for theme in expertise_themes:
        for s in re.split(r'(?<=[.!?])\s+', theme):
            if len(s.split()) >= 8:
                theme_sentences.append(s.strip())
            if len(theme_sentences) >= 1:
                break
        if len(theme_sentences) >= 1:
            break

    # Compose the explanation
    summary_parts = []
    summary_parts.append(f"According to our analysis, the expertise of {expertise_name} is a {match_quality} for this call (confidence {confidence:.0%}).")
    if contributors:
        summary_parts.append(f"Top contributors to this match: {', '.join(contributors)}.")

    if cap_sentences:
        summary_parts.append(cap_sentences[0])
    if theme_sentences:
        summary_parts.append(theme_sentences[0])

    # Add a domain overlap sentence if present
    if shared_domains:
        doms = ', '.join(shared_domains[:2])
        summary_parts.append(f"There is notable overlap in the following domains: {doms}.")

    # Add a problem-solution match sentence if present
    direct_matches = content_relevance.get("direct_matches", 0)
    prob_count = content_relevance.get("problem_count", 0)
    sol_count = content_relevance.get("solution_count", 0)
    if direct_matches > 0:
        summary_parts.append(f"The expertise directly addresses key requirements with {direct_matches} strong problem-solution matches.")
    elif prob_count > 0 and sol_count > 0:
        summary_parts.append(f"The expertise addresses the call's challenges with {sol_count} relevant solutions.")

    # Add a theme alignment sentence if present
    if theme_sentences:
        summary_parts.append(f"This aligns with the strategic themes of the expertise.")

    # Add TRL alignment
    trl_match = thematic_match.get("trl_match")
    if trl_match is not None:
        if trl_match:
            summary_parts.append("The technology readiness level is well aligned with the call's requirements.")
        else:
            summary_parts.append("The technology readiness level may differ, so project stage compatibility should be checked.")

    # Add a final recommendation
    if confidence >= 0.55:
        rec = "This is a high-priority fit and could justify a leadership or major work-package role."
    elif confidence >= 0.45:
        rec = "This is a strong match and suitable for a core work-package contribution."
    elif confidence >= 0.35:
        rec = "This is a good potential fit; consider a focused role tied to specific deliverables."
    else:
        rec = "This is a possible fit; review detailed call requirements to target niche contributions."
    summary_parts.append(rec)

    # Join all parts into a single paragraph
    return ' '.join([p for p in summary_parts if p])

def is_valid_description(description):
    """Check if description is meaningful."""
    if pd.isna(description):
        return False
    desc_str = str(description).strip().lower()
    # Accept 'tbd' as valid (do not skip)
    return bool(desc_str) and (desc_str == "tbd" or len(desc_str) > 20)

def analyze_expertise_files(expertise_dir):
    """
    Analyze all expertise files in a directory.
    Returns a dictionary of expertise info by expertise name.
    """
    expertise_files = sorted(list(Path(expertise_dir).glob("*.txt")))
    all_expertise = {}
    
    for expertise_file in expertise_files:
        expertise_info = extract_expertise_description(expertise_file)
        all_expertise[expertise_info["name"]] = expertise_info
    
    return all_expertise

def main(expertise_dir=None):
    # Main analysis
    print("=" * 110)
    print("QUALITATIVE EXPERTISE ANALYSIS - VALIDATION-BASED MATCHING".center(110))
    print("=" * 110)
    print()

    # Discover expertise files
    print("Discovering expertise files...\n")
    if expertise_dir is None:
        expertise_dir = Path(__file__).parent.parent.parent / "Expertises"
    else:
        expertise_dir = Path(expertise_dir)
    expertise_files = sorted(list(expertise_dir.glob("*.txt")))

    if not expertise_files:
        print("[WARNING] No expertise files found in Expertises/ folder")
        exit(1)

    print(f"Found {len(expertise_files)} expertise file(s):")
    for f in expertise_files:
        print(f"  - {f.name}")
    print()

    # Extract expertise information
    all_expertise = {}
    for expertise_file in expertise_files:
        expertise_info = extract_expertise_description(expertise_file)
        all_expertise[expertise_info["name"]] = expertise_info
        print(f"[OK] Analyzed {expertise_info['name']}: {expertise_info['content_length']} chars, {len(expertise_info['themes'])} themes, {len(expertise_info['capabilities'])} capabilities")

    print()

    # Read Excel file
    excel_file = Path(__file__).parent.parent / "Horizon_Europe_Funding_Calls_2026-2027_Complete_final.xlsx"
    excel_file = str(excel_file)
    
    try:
        df = pd.read_excel(excel_file, engine='openpyxl')
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        print(f"Trying alternative Excel file...")
        excel_file_alt = Path(__file__).parent.parent / "Horizon_Calls_583_calls.xlsx"
        excel_file_alt = str(excel_file_alt)
        df = pd.read_excel(excel_file_alt, engine='openpyxl')

    # Remove old analysis columns
    cols_to_remove = [col for col in df.columns if col in ["Touchpoints", "Touchpoint expertise"] or "Why & How" in col or "Validation" in col]
    df = df.drop(cols_to_remove, axis=1)

    # PRE-COMPUTE EMBEDDINGS FOR SPEED (cache all texts upfront)
    print("Pre-computing embeddings for speed optimization...")
    texts_to_embed = []
    
    # Collect all call texts
    for idx, row in df.iterrows():
        title = str(row['Title'])
        description = str(row['Description'])
        if is_valid_description(description):
            call_text = description
        else:
            call_text = ""
        combined = f"{title} {call_text}".strip()
        if combined:
            texts_to_embed.append(combined)
    
    # Collect all expertise topic texts and projects
    for expertise_name, expertise_info in all_expertise.items():
        # Add themes and capabilities as separate embedding texts
        for theme in expertise_info.get("themes", [])[:3]:  # Top 3 themes
            if theme:
                texts_to_embed.append(theme)
        for cap in expertise_info.get("capabilities", [])[:2]:  # Top 2 capabilities
            if cap:
                texts_to_embed.append(cap)
        
        # Add project descriptions
        projects = expertise_info.get("projects", [])
        for proj in projects[:3]:  # Top 3 projects
            if proj:
                texts_to_embed.append(proj)
    
    # Batch encode all texts (much faster than individual encoding)
    if texts_to_embed:
        print(f"  Embedding {len(texts_to_embed)} texts in batch...")
        batch_encode_texts(texts_to_embed)
        print(f"  [OK] Cache populated with {len(_embedding_cache)} embeddings")
    
    print()

    # Analyze each call
    matched_calls_data = []
    calls_with_desc = 0
    calls_with_matches = 0

    print("Running validation-based analysis...")
    print()

    for idx, row in df.iterrows():
        call_id = row['Call']
        title = str(row['Title'])
        description = str(row['Description'])
        
        # Use description if valid, otherwise use empty string
        if is_valid_description(description):
            call_text = description
        else:
            call_text = ""  # Empty string - title is already passed separately
        
        # Skip if title is also invalid
        if not title or len(title.strip()) < 5:
            continue
        
        calls_with_desc += 1
        
        # Test each expertise against this call
        call_matches = []
        
        for expertise_name, expertise_info in all_expertise.items():
            # Run comprehensive validation using the available text
            validation = validate_alignment_comprehensive(title, call_text, expertise_info)
            
            # Store match if it passes validation
            if validation["overall_pass"]:
                call_matches.append({
                    "expertise": expertise_name,
                    "expertise_info": expertise_info,
                    "validation": validation
                })
        
        if call_matches:
            calls_with_matches += 1
            # Sort by confidence
            call_matches.sort(key=lambda x: x["validation"]["confidence"], reverse=True)
            
            matched_calls_data.append({
                "index": idx,
                "call_id": call_id,
                "title": title,
                "description": description,
                "matches": call_matches
            })

    print(f"Analysis complete: {calls_with_matches}/{calls_with_desc} calls identified solid expertise matches\n")

    # Add columns
    df["Touchpoints"] = ""

    # Add validation-based explanation columns for each expertise
    matched_expertises = set()
    for call_data in matched_calls_data:
        for match in call_data["matches"]:
            matched_expertises.add(match["expertise"])

    for expertise in sorted(matched_expertises):
        # Only add Confidence column, never create Project Alignment
        if f"{expertise} - Confidence" not in df.columns:
            df[f"{expertise} - Confidence"] = ""

    # Populate data
    for call_data in matched_calls_data:
        idx = call_data["index"]
        # Set touchpoints
        touchpoints = "; ".join(m["expertise"] for m in call_data["matches"])
        df.at[idx, "Touchpoints"] = touchpoints
        # Add only confidence
        for match in call_data["matches"]:
            expertise = match["expertise"]
            col_confidence = f"{expertise} - Confidence"
            confidence = match["validation"]["confidence"]
            df.at[idx, col_confidence] = f"{confidence:.1%}"

    # Reorder columns to put Touchpoints next to Description
    cols = df.columns.tolist()
    if 'Description' in cols and 'Touchpoints' in cols:
        desc_idx = cols.index('Description')
        touch_idx = cols.index('Touchpoints')
        # Remove Touchpoints from its current position
        cols.pop(touch_idx)
        # Insert it right after Description
        cols.insert(desc_idx + 1, 'Touchpoints')
        df = df[cols]

    # Save
    print(f"Saving Excel with qualitative analysis...")
    df.to_excel(excel_file, index=False, engine='openpyxl')

    # Apply formatting
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.freeze_panes = "A2"

    light_gray = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = light_gray if (row_idx - 2) % 2 == 0 else white
        for cell in row:
            cell.border = border
            cell.fill = fill
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            if cell.column == 10 and cell.value:
                if isinstance(cell.value, str) and cell.value.startswith('http'):
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0563C1", underline="single")

    wb.save(excel_file)

    # Print report (simplified to avoid encoding issues)
    print("\n" + "="*110)
    print("QUALITATIVE MATCHING REPORT - VALIDATION-BASED EXPERTISE ALIGNMENT".center(110))
    print("="*110 + "\n")

    report_count = 0
    for i, call_data in enumerate(matched_calls_data, 1):
        try:
            # Safely encode/decode to avoid Unicode issues
            call_id = call_data['call_id'].encode('ascii', errors='replace').decode()
            print(f"{i}. {call_id}")
            print()
            report_count += 1
            
            if report_count >= 50:  # Limit report to first 50 calls to avoid encoding issues
                break
        except Exception as e:
            continue

    print("="*110)
    print(f"Summary: {calls_with_matches} calls with validated expertise matches | {len(matched_expertises)} expertise areas engaged")
    print("Validation approach: Continuous relevance scoring (semantic + domain + project history)")
    print(f"Excel updated: {excel_file}")
    print("="*110 + "\n")


if __name__ == "__main__":
    main()
